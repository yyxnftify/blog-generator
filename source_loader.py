"""
Source Loader モジュール
PDF / Excel / テキスト / 画像 / インスタ貼り付け など
さまざまな情報ソースを一元管理し、記事生成に渡すデータを準備する。

ローカル（blog_data/sources/）とクラウド（Google Sheets）の
ハイブリッド方式に対応。クラウド接続時はGoogle Sheetsを優先する。
"""

import os
import json
import glob
from datetime import datetime

# ==========================================
# 設定
# ==========================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCES_DIR = os.path.join(BASE_DIR, "blog_data", "sources")
INSTAGRAM_FILE = os.path.join(BASE_DIR, "blog_data", "instagram_sources.json")

# ソースフォルダがなければ作成
os.makedirs(SOURCES_DIR, exist_ok=True)

# 対応する拡張子
TEXT_EXTENSIONS = [".txt", ".md", ".csv"]
PDF_EXTENSIONS = [".pdf"]
EXCEL_EXTENSIONS = [".xlsx", ".xls"]
IMAGE_EXTENSIONS = [".jpg", ".jpeg", ".png", ".gif", ".webp"]

# クラウドモード判定用（遅延import）
_cloud_module = None


def _get_cloud():
    """blog_sheet_managerモジュールを遅延読み込みし、接続可能か確認する"""
    global _cloud_module
    if _cloud_module is None:
        try:
            import blog_sheet_manager
            _cloud_module = blog_sheet_manager
        except ImportError:
            _cloud_module = False  # インポート失敗
    
    if _cloud_module is False:
        return None
    
    try:
        if _cloud_module.is_connected():
            return _cloud_module
    except Exception:
        pass
    return None


# ==========================================
# テキストファイル読み込み
# ==========================================

def load_text_file(filepath, max_chars=50000):
    """テキストファイルを読み込む"""
    try:
        # いくつかのエンコーディングを試す
        for encoding in ["utf-8", "utf-8-sig", "cp932", "shift_jis", "euc-jp"]:
            try:
                with open(filepath, "r", encoding=encoding) as f:
                    content = f.read()
                if content:
                    if len(content) > max_chars:
                        content = content[:max_chars] + "\n...(以下省略)"
                    return {
                        "type": "text",
                        "filename": os.path.basename(filepath),
                        "content": content,
                        "char_count": len(content),
                    }
            except (UnicodeDecodeError, UnicodeError):
                continue
        return None
    except Exception as e:
        print(f"テキスト読み込みエラー ({filepath}): {e}")
        return None


# ==========================================
# PDF読み込み
# ==========================================

def load_pdf_file(filepath, max_chars=50000):
    """PDFファイルからテキストを抽出する"""
    try:
        import pdfplumber
    except ImportError:
        print("⚠ pdfplumber がインストールされていません。pip install pdfplumber を実行してください。")
        # フォールバック: PyPDF2を試す
        try:
            from PyPDF2 import PdfReader
            return _load_pdf_pypdf2(filepath, max_chars)
        except ImportError:
            print("⚠ PyPDF2 も見つかりません。PDF読み込みにはいずれかが必要です。")
            return None

    try:
        text_parts = []
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(f"[ページ{i+1}]\n{page_text}")

                # テーブルも抽出
                tables = page.extract_tables()
                for table in tables:
                    if table:
                        table_text = _format_table(table)
                        if table_text:
                            text_parts.append(f"[ページ{i+1} 表]\n{table_text}")

        content = "\n\n".join(text_parts)
        if len(content) > max_chars:
            content = content[:max_chars] + "\n...(以下省略)"

        return {
            "type": "pdf",
            "filename": os.path.basename(filepath),
            "content": content,
            "char_count": len(content),
            "page_count": len(text_parts),
        }
    except Exception as e:
        print(f"PDF読み込みエラー ({filepath}): {e}")
        return None


def _load_pdf_pypdf2(filepath, max_chars=50000):
    """PyPDF2でのフォールバック読み込み"""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(filepath)
        text_parts = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text:
                text_parts.append(f"[ページ{i+1}]\n{text}")

        content = "\n\n".join(text_parts)
        if len(content) > max_chars:
            content = content[:max_chars] + "\n...(以下省略)"

        return {
            "type": "pdf",
            "filename": os.path.basename(filepath),
            "content": content,
            "char_count": len(content),
            "page_count": len(text_parts),
        }
    except Exception as e:
        print(f"PyPDF2読み込みエラー ({filepath}): {e}")
        return None


def _format_table(table):
    """表データをテキスト形式に変換する"""
    if not table:
        return ""
    lines = []
    for row in table:
        if row:
            cells = [str(cell).strip() if cell else "" for cell in row]
            lines.append(" | ".join(cells))
    return "\n".join(lines)


# ==========================================
# Excel読み込み
# ==========================================

def load_excel_file(filepath, max_chars=50000):
    """Excelファイルからデータを抽出する"""
    try:
        import openpyxl
    except ImportError:
        print("⚠ openpyxl がインストールされていません。pip install openpyxl を実行してください。")
        return None

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines = [f"【シート: {sheet_name}】"]

            for row in ws.iter_rows(values_only=True):
                cells = [str(cell).strip() if cell is not None else "" for cell in row]
                # 空行はスキップ
                if any(c for c in cells):
                    sheet_lines.append(" | ".join(cells))

            if len(sheet_lines) > 1:  # ヘッダだけじゃない場合
                text_parts.append("\n".join(sheet_lines))

        wb.close()

        content = "\n\n".join(text_parts)
        if len(content) > max_chars:
            content = content[:max_chars] + "\n...(以下省略)"

        return {
            "type": "excel",
            "filename": os.path.basename(filepath),
            "content": content,
            "char_count": len(content),
            "sheet_count": len(wb.sheetnames),
        }
    except Exception as e:
        print(f"Excel読み込みエラー ({filepath}): {e}")
        return None


# ==========================================
# 画像ファイルの管理
# ==========================================

def load_image_info(filepath):
    """画像ファイルの情報を取得する（テキスト抽出はしない、メタ情報のみ）"""
    try:
        filename = os.path.basename(filepath)
        file_size = os.path.getsize(filepath)

        # ファイル名から説明を推測（日本語ファイル名をそのまま活用）
        name_without_ext = os.path.splitext(filename)[0]

        return {
            "type": "image",
            "filename": filename,
            "filepath": filepath,
            "content": f"[画像ファイル] {filename} (サイズ: {file_size // 1024}KB)",
            "description": name_without_ext,
            "char_count": 0,
        }
    except Exception as e:
        print(f"画像情報取得エラー ({filepath}): {e}")
        return None


# ==========================================
# Instagram投稿ソースの管理
# ==========================================

def load_instagram_sources():
    """保存済みのInstagram投稿ソースを読み込む"""
    if not os.path.exists(INSTAGRAM_FILE):
        return []
    try:
        with open(INSTAGRAM_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception as e:
        print(f"Instagramソース読み込みエラー: {e}")
        return []


def save_instagram_source(account_name, caption_text, post_url="", tags=""):
    """
    Instagram投稿のキャプションをソースとして保存する。
    クラウド接続時はGoogle Sheetsに、ローカル時はJSONファイルに保存。
    """
    cloud = _get_cloud()
    if cloud:
        # クラウド保存
        return cloud.add_instagram(account_name, caption_text, post_url, tags)
    
    # ローカル保存（フォールバック）
    sources = load_instagram_sources()

    new_entry = {
        "id": len(sources) + 1,
        "account_name": account_name.strip(),
        "caption": caption_text.strip(),
        "post_url": post_url.strip(),
        "tags": tags.strip(),
        "saved_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    sources.append(new_entry)

    try:
        with open(INSTAGRAM_FILE, "w", encoding="utf-8") as f:
            json.dump(sources, f, ensure_ascii=False, indent=2)
        print(f"✅ Instagramソース保存: @{account_name} ({len(caption_text)}文字)")
        return True
    except Exception as e:
        print(f"Instagramソース保存エラー: {e}")
        return False


def delete_instagram_source(source_id):
    """指定IDのInstagramソースを削除する"""
    cloud = _get_cloud()
    if cloud:
        return cloud.delete_instagram(source_id)
    
    # ローカル削除
    sources = load_instagram_sources()
    sources = [s for s in sources if s.get("id") != source_id]

    for i, s in enumerate(sources):
        s["id"] = i + 1

    try:
        with open(INSTAGRAM_FILE, "w", encoding="utf-8") as f:
            json.dump(sources, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


def get_instagram_text_for_keyword(keyword=""):
    """
    キーワードに関連するInstagramソースのテキストを統合して返す。
    キーワードが空の場合は全ソースを返す。
    """
    sources = load_instagram_sources()
    if not sources:
        return ""

    relevant = []
    for s in sources:
        if not keyword:
            relevant.append(s)
        else:
            # キーワードがタグやキャプションに含まれているか
            search_text = (s.get("caption", "") + s.get("tags", "")).lower()
            if any(kw.lower() in search_text for kw in keyword.split()):
                relevant.append(s)

    if not relevant:
        # 関連するものがなければ全部返す（情報量を確保）
        relevant = sources

    parts = []
    for s in relevant:
        part = f"【Instagram @{s['account_name']}】\n{s['caption']}"
        if s.get("post_url"):
            part += f"\n(出典: {s['post_url']})"
        parts.append(part)

    return "\n\n---\n\n".join(parts)


# ==========================================
# フォルダ内ソースの一括読み込み
# ==========================================

def load_all_file_sources(sources_dir=None):
    """
    sources/ フォルダ内の全ファイルを読み込んで統合する。

    Returns:
        dict: {
            "text_sources": list,
            "pdf_sources": list,
            "excel_sources": list,
            "image_sources": list,
            "total_count": int,
            "combined_text": str,
        }
    """
    if sources_dir is None:
        sources_dir = SOURCES_DIR

    result = {
        "text_sources": [],
        "pdf_sources": [],
        "excel_sources": [],
        "image_sources": [],
        "total_count": 0,
        "combined_text": "",
    }

    if not os.path.exists(sources_dir):
        return result

    all_files = []
    for f in os.listdir(sources_dir):
        filepath = os.path.join(sources_dir, f)
        if os.path.isfile(filepath):
            all_files.append(filepath)

    text_parts = []

    for filepath in sorted(all_files):
        ext = os.path.splitext(filepath)[1].lower()

        if ext in TEXT_EXTENSIONS:
            data = load_text_file(filepath)
            if data:
                result["text_sources"].append(data)
                text_parts.append(f"【ファイル: {data['filename']}】\n{data['content']}")

        elif ext in PDF_EXTENSIONS:
            data = load_pdf_file(filepath)
            if data:
                result["pdf_sources"].append(data)
                text_parts.append(f"【PDF: {data['filename']}】\n{data['content']}")

        elif ext in EXCEL_EXTENSIONS:
            data = load_excel_file(filepath)
            if data:
                result["excel_sources"].append(data)
                text_parts.append(f"【Excel: {data['filename']}】\n{data['content']}")

        elif ext in IMAGE_EXTENSIONS:
            data = load_image_info(filepath)
            if data:
                result["image_sources"].append(data)
                # 画像はテキストとしては追加しない（メタ情報のみ）

    result["total_count"] = (
        len(result["text_sources"]) +
        len(result["pdf_sources"]) +
        len(result["excel_sources"]) +
        len(result["image_sources"])
    )

    result["combined_text"] = "\n\n===\n\n".join(text_parts)

    return result


def get_all_sources_text(keyword=""):
    """
    全ソース（ローカルファイル + クラウド + Instagram）を統合してテキストとして返す。
    クラウド接続時はGoogle Sheetsのデータも含める。
    """
    parts = []

    cloud = _get_cloud()

    if cloud:
        # クラウドモード: Google Sheetsからデータ取得
        cloud_text = cloud.get_all_cloud_sources_text(keyword)
        if cloud_text:
            parts.append(cloud_text)
    else:
        # ローカルモード: ファイルシステムからデータ取得
        file_sources = load_all_file_sources()
        if file_sources["combined_text"]:
            parts.append("## ★ ファイルソース（独自情報）\n" + file_sources["combined_text"])

        insta_text = get_instagram_text_for_keyword(keyword)
        if insta_text:
            parts.append("## ★ Instagramソース（専門家投稿）\n" + insta_text)

    # Webページソース（ローカル・クラウド共通）
    web_text = get_web_sources_text(keyword)
    if web_text:
        parts.append("## ★ Webページソース（参考URL）\n" + web_text)

    combined = "\n\n" + "=" * 40 + "\n\n".join(parts)

    if len(combined) > 80000:
        combined = combined[:80000] + "\n...(以下省略)"

    return combined


def save_uploaded_file(uploaded_file, target_dir=None):
    """
    アップロードファイルを保存する。
    クラウド接続時: ファイルをテキスト化してGoogle Sheetsに保存
    ローカル時: ファイルシステムに保存
    """
    cloud = _get_cloud()
    
    # まずローカルに一時保存（PDF/Excel読み込みのため）
    if target_dir is None:
        target_dir = SOURCES_DIR
    os.makedirs(target_dir, exist_ok=True)
    
    try:
        filepath = os.path.join(target_dir, uploaded_file.name)
        with open(filepath, "wb") as f:
            f.write(uploaded_file.getbuffer())
    except Exception as e:
        print(f"一時ファイル保存エラー: {e}")
        return None
    
    # クラウド対応: テキスト抽出してGoogle Sheetsに保存
    if cloud:
        ext = os.path.splitext(uploaded_file.name)[1].lower()
        content = ""
        file_type = "text"
        
        if ext in TEXT_EXTENSIONS:
            data = load_text_file(filepath)
            if data:
                content = data["content"]
                file_type = "text"
        elif ext in PDF_EXTENSIONS:
            data = load_pdf_file(filepath)
            if data:
                content = data["content"]
                file_type = "pdf"
        elif ext in EXCEL_EXTENSIONS:
            data = load_excel_file(filepath)
            if data:
                content = data["content"]
                file_type = "excel"
        elif ext in IMAGE_EXTENSIONS:
            content = f"[画像ファイル] {uploaded_file.name}"
            file_type = "image"
        
        if content:
            cloud.add_source(uploaded_file.name, file_type, content)
    
    print(f"✅ ファイル保存: {filepath}")
    return filepath


def get_source_summary():
    """
    現在のソース状況のサマリーを返す。
    クラウド接続時はGoogle Sheetsからも集計。
    """
    cloud = _get_cloud()
    
    if cloud:
        return cloud.get_cloud_source_summary()
    
    # ローカルモード
    file_sources = load_all_file_sources()
    insta_sources = load_instagram_sources()
    web_sources = load_web_sources()

    return {
        "text_count": len(file_sources["text_sources"]),
        "pdf_count": len(file_sources["pdf_sources"]),
        "excel_count": len(file_sources["excel_sources"]),
        "image_count": len(file_sources["image_sources"]),
        "instagram_count": len(insta_sources),
        "web_count": len(web_sources),
        "total_file_count": file_sources["total_count"],
        "total_count": file_sources["total_count"] + len(insta_sources) + len(web_sources),
    }


# ==========================================
# Webページソースの管理
# ==========================================

WEB_SOURCES_FILE = os.path.join(BASE_DIR, "blog_data", "web_sources.json")


def fetch_web_page(url, max_chars=30000):
    """
    URLからWebページのテキスト内容を取得する。
    HTMLタグを除去して本文テキストのみ抽出する。
    """
    try:
        import requests
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        
        # エンコーディング自動判定
        response.encoding = response.apparent_encoding or "utf-8"
        html = response.text
        
        # BeautifulSoupでテキスト抽出
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(html, "html.parser")
            
            # 不要なタグを除去
            for tag in soup(["script", "style", "nav", "footer", "header", "aside", "form"]):
                tag.decompose()
            
            # ページタイトル取得
            title = soup.title.string.strip() if soup.title and soup.title.string else url
            
            # 本文テキスト抽出
            text = soup.get_text(separator="\n", strip=True)
            
            # 空行の重複を除去
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            text = "\n".join(lines)
            
        except ImportError:
            # BeautifulSoupがない場合はHTML正規表現で簡易除去
            import re
            title = url
            text = re.sub(r'<[^>]+>', ' ', html)
            text = re.sub(r'\s+', ' ', text).strip()
        
        # 文字数制限
        if len(text) > max_chars:
            text = text[:max_chars] + "\n...(以下省略)"
        
        return {
            "success": True,
            "title": title,
            "text": text,
            "char_count": len(text),
            "url": url
        }
        
    except Exception as e:
        return {
            "success": False,
            "title": "",
            "text": "",
            "char_count": 0,
            "url": url,
            "error": str(e)
        }


def load_web_sources():
    """保存済みのWebページソースを読み込む"""
    if os.path.exists(WEB_SOURCES_FILE):
        try:
            with open(WEB_SOURCES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return []
    return []


def save_web_source(url, title="", text="", tags=""):
    """
    Webページの内容をソースとして保存する。
    クラウド接続時はGoogle Sheetsにも保存。
    """
    sources = load_web_sources()
    
    new_source = {
        "id": datetime.now().strftime('%Y%m%d_%H%M%S') + f"_{len(sources)}",
        "url": url,
        "title": title,
        "content": text,
        "char_count": len(text),
        "tags": tags,
        "saved_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    sources.append(new_source)
    
    try:
        os.makedirs(os.path.dirname(WEB_SOURCES_FILE), exist_ok=True)
        with open(WEB_SOURCES_FILE, "w", encoding="utf-8") as f:
            json.dump(sources, f, ensure_ascii=False, indent=2)
        
        # クラウドにも保存
        cloud = _get_cloud()
        if cloud:
            try:
                cloud.save_source("web", url, text[:10000], title)
            except:
                pass
        
        return True
    except Exception as e:
        print(f"Webソース保存エラー: {e}")
        return False


def delete_web_source(source_id):
    """指定IDのWebソースを削除する"""
    sources = load_web_sources()
    sources = [s for s in sources if s.get("id") != source_id]
    try:
        with open(WEB_SOURCES_FILE, "w", encoding="utf-8") as f:
            json.dump(sources, f, ensure_ascii=False, indent=2)
        return True
    except:
        return False


def get_web_sources_text(keyword=""):
    """
    保存済みWebソースのテキストを統合して返す。
    キーワードでフィルタリング可能。
    """
    sources = load_web_sources()
    if not sources:
        return ""
    
    parts = []
    for src in sources:
        # キーワードマッチング
        if keyword:
            content_lower = (src.get("content", "") + src.get("tags", "") + src.get("title", "")).lower()
            kw_lower = keyword.lower()
            # キーワードに部分一致するもの、またはキーワードなしなら全件
            if kw_lower not in content_lower and all(k not in content_lower for k in kw_lower.split()):
                continue
        
        title = src.get("title", src.get("url", ""))
        content = src.get("content", "")[:5000]
        parts.append(f"### Webソース: {title}\nURL: {src.get('url', '')}\n{content}")
    
    return "\n\n".join(parts)


def is_cloud_mode():
    """現在クラウドモードかどうかを返す"""
    return _get_cloud() is not None


# テスト用
if __name__ == "__main__":
    print("=== ソースローダー テスト ===")
    print(f"クラウドモード: {is_cloud_mode()}")
    summary = get_source_summary()
    print(f"テキスト: {summary['text_count']}件")
    print(f"PDF: {summary['pdf_count']}件")
    print(f"Excel: {summary['excel_count']}件")
    print(f"画像: {summary['image_count']}件")
    print(f"Instagram: {summary['instagram_count']}件")
    print(f"合計: {summary['total_count']}件")
